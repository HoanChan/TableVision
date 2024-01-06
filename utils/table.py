import xml.etree.ElementTree as ET
import base64

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

def cells_to_html(cells):

    table = ET.Element("table")
    current_row = -1

    for cell in cells:
        this_row = cell['row']

        attrib = {}
        colspan = cell['col_span']
        if colspan > 1:
            attrib['colspan'] = str(colspan)
        rowspan = cell['row_span']
        if rowspan > 1:
            attrib['rowspan'] = str(rowspan)
        if this_row > current_row:
            current_row = this_row
            if current_row == 0:
                cell_tag = "th"
                row = ET.SubElement(table, "thead")
            else:
                cell_tag = "td"
                row = ET.SubElement(table, "tr")
        tcell = ET.SubElement(row, cell_tag, attrib=attrib)
        tcell.text = cell['cell text']

    return str(ET.tostring(table, encoding="unicode", short_empty_elements=False))

def image_to_base64(image_path):
    with open(image_path, "rb") as image_file:
        encoded_string = base64.b64encode(image_file.read()).decode("utf-8")
    return encoded_string

def createHTML(image_path, html, show_image = True, useBase64 = False):
    css = """
    table {
    border-collapse: collapse;
    border-spacing: 0;
    width: 100%;
    font-family: sans-serif;
    }

    th, td {
    padding: 8px;
    text-align: center;
    vertical-align: middle;
    }

    th {
    font-weight: bold;
    font-size: 1.2em;
    }

    tr:nth-child(even) {
    background-color: rgba(68, 68, 68, 0.2);
    }

    table th, table td {
    border: 1px solid rgba(68, 68, 68, 0.5);
    }
    """
    image_tag = """<img src='""" + image_path + """' alt="Ảnh link" style="width: 90%;">"""
    if useBase64:
      image_tag = '<img src="data:image/jpeg;base64,' + image_to_base64(image_path) + '" alt="Ảnh Base64" style="width: 90%;">'
    body = """
    <div style="display: flex;">
    <div style="flex: 1;">
    """ + image_tag + """
    </div>
    <div style="flex: 1;">""" + html + """</div>
    </div>
    """ if show_image else html
    new_html = '<head><style>'+ css +'</style></head><body>'+ body + '</body>'
    return new_html

def cells_to_xlsx(cells, path = "data.xlsx"):
    # Tạo một Workbook mới
    wb = Workbook()
    sheet = wb.active

    # Đặt giá trị và gộp ô cho mỗi phần tử trong data
    for cell in cells:
        row = cell['row'] + 1
        column = cell['col'] + 1
        rowspan = cell['row_span']
        colspan = cell['col_span']
        text = cell['cell text']
        print(f'row: {row}, column: {column}, rowspan: {rowspan}, colspan: {colspan}')

        # Đặt giá trị cho ô
        top_left_cell = sheet.cell(row=row, column=column)
        top_left_cell.value = text
        top_left_cell.alignment = Alignment(horizontal='center',    # Căn giữa theo chiều ngang
                                            vertical='center',      # Căn giữa theo chiều dọc
                                            wrap_text=True)         # Tự động xuống dòng khi text quá dài

        # Gộp ô nếu rowspan hoặc colspan lớn hơn 1
        if rowspan > 1 or colspan > 1:
            merge_range = '{}{}:{}{}'.format(
                get_column_letter(column),
                row,
                get_column_letter(column + colspan - 1),
                row + rowspan - 1
            )
            sheet.merge_cells(merge_range)
            print(f'Merged {merge_range}')

    wb.save(path)
    print(f'Saved to {path}')